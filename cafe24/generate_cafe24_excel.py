# -*- coding: utf-8 -*-
"""
현담토목 품목 단가(data/품목.csv)를 카페24 '상품 등록용 엑셀' 형식으로 변환하는 스크립트.

실행:  python3 cafe24/generate_cafe24_excel.py        (저장소 루트에서)
결과:  cafe24/현담토목_카페24_상품등록.xlsx

가격 원본은 오직 data/품목.csv 입니다. (판매단가 = 원가 × (1+마진율), 부가세 별도)
가격을 바꾸려면 data/품목.csv 를 고치고(또는 ops.py 판매가설정) 이 스크립트를 다시 실행하면 됩니다.

시트 구성
  1) 카페24_규격별등록  : 품목 하나 = 상품 하나 (권장)
  2) 카페24_옵션형등록  : 제품군 하나 = 상품 하나 + 규격 옵션 (업로드 후 옵션별 추가금액 입력 필요)
  3) 옵션추가금액표     : 2)번 시트용 규격별 추가금액
  4) 제품마스터_단가표  : 사람이 보는 단가표 (부가세 별도/포함, 나라장터 식별번호)
  5) 카페24_열설명      : 각 열에 무엇을 넣었는지 설명
"""
import csv
import html
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
ITEMS_CSV = os.path.join(ROOT, "data", "품목.csv")
OUT_XLSX = os.path.join(HERE, "현담토목_카페24_상품등록.xlsx")

# 카페24 판매가에 부가세를 포함할지. True: 판매가 = 판매단가 × 1.1 (10원 단위 반올림), 공급가 = 판매단가(부가세 별도)
VAT_INCLUDED = True

COMPANY = "주식회사 현담토목"
PHONE = "010-9999-1335"
FAX = "061-362-3621"
EMAIL = "khddddd@naver.com"
ADDR = "광주광역시 광산구 산월동 868-11"
HOMEPAGE = "https://hyundamtomok.wordpress.com"
LOGO = "https://hyundamtomok.wordpress.com/wp-content/uploads/2026/09/hyundam-logo.png"

# 카페24 '상품 등록용 엑셀' 열 순서 (2024~2025 양식 기준, AG=옵션사용 ~ AO=필수여부).
CAFE24_COLUMNS = [
    "상품코드", "자체 상품코드", "진열상태", "판매상태", "상품분류 번호", "상품분류 신상품영역", "상품분류 추천상품영역",
    "상품명", "영문 상품명", "상품명(관리용)", "공급사 상품명", "모델명", "상품 요약설명", "상품 간략설명", "상품 상세설명",
    "모바일 상품 상세설명 설정", "모바일 상품 상세설명", "검색어설정", "과세구분", "소비자가", "공급가", "상품가", "판매가",
    "판매가 대체문구 사용", "판매가 대체문구", "주문수량 제한 기준", "최소 주문수량(이상)", "최대 주문수량(이하)",
    "적립금", "적립금 구분", "공통이벤트 정보", "성인인증",
    "옵션사용", "옵션구성방식", "옵션 표시방식", "옵션세트명", "옵션 입력", "옵션 스타일", "버튼이미지 설정", "색상 설정", "필수여부",
    "품절표시 문구", "추가입력옵션", "추가입력옵션 명칭", "추가입력옵션 선택/필수여부", "입력글자수(자)",
    "이미지등록(상세)", "이미지등록(목록)", "이미지등록(작은목록)", "이미지등록(축소)", "이미지등록(추가)",
    "제조사", "공급사", "브랜드", "트렌드", "자체분류 코드", "원산지", "상품부피(cm)", "상품소재", "영문 상품소재",
    "상품결제안내", "상품배송안내", "교환/반품안내", "서비스문의/안내",
    "배송정보", "배송방법", "국내/해외배송", "배송지역", "배송비 선결제 설정", "배송기간", "배송비 구분", "배송비입력", "스토어픽업 설정",
    "상품 전체중량(kg)", "HS코드", "상품 구분(해외통관)", "상품소재(해외통관)", "영문 상품소재(해외통관)", "옷감(해외통관)",
    "검색엔진 노출 설정", "검색엔진 노출 제목", "검색엔진 노출 작성자", "검색엔진 노출 설명", "검색엔진 노출 키워드",
    "개별결제수단설정", "상품 유효기간 사용", "상품 유효기간",
]

PAYMENT_GUIDE = ("무통장입금·카드결제 모두 가능합니다. 세금계산서가 필요하시면 주문 시 요청사항에 사업자번호와 이메일을 적어 주세요. "
                 "결제 확인 후 출고(입금 확인 후 상차)하며, 대량 주문은 전화 견적 후 계약서 기준으로 진행합니다.")
SHIPPING_GUIDE = ("콘크리트·주철 제품은 무게가 많이 나가 택배가 불가능하며 화물차(카고·크레인)로 배송합니다. 운반비는 별도(착불)이며 "
                  "거리·물량·차량 종류(5톤/11톤/크레인)에 따라 달라집니다. 광주·전남 현장은 당일~2일, 그 외 지역은 2~4일 소요됩니다. "
                  "현장 진입로에 대형차가 들어갈 수 있어야 하며, 하차는 현장 부담입니다(크레인·인력이 필요하면 별도 협의). "
                  f"문의 {PHONE}")
RETURN_GUIDE = ("제품 특성상 단순 변심 반품은 어렵습니다. 파손·규격 오배송은 하차 전 확인 후 즉시 연락 주시면 교환해 드립니다. "
                "하차 완료 후에는 파손 책임을 판단하기 어려우니 인수 시 반드시 확인해 주세요.")
SERVICE_GUIDE = f"규격 선택·수량 계산·시공 방법 상담: 전화 {PHONE} (평일 08:00~18:00) / 팩스 {FAX} / 이메일 {EMAIL}"

# 카페24 분류(자체분류 코드) — 관리자 > 분류 관리에서 이 5개를 만들면 됩니다.
CATEGORIES = [
    ("C01", "콘크리트 수로관"),
    ("C02", "콘크리트 맨홀·집수정"),
    ("C03", "주철 맨홀뚜껑·인상링·트렌치"),
    ("C04", "PE관·부속"),
    ("C05", "루프드레인"),
]
CAT_NAME = dict(CATEGORIES)


def category_of(대분류):
    if 대분류.startswith("콘크리트-집수정") or 대분류.startswith("콘크리트-맨홀"):
        return "C02"
    if 대분류.startswith("콘크리트"):
        return "C01"
    if 대분류.startswith("주철"):
        return "C03"
    if 대분류.startswith("PE관"):
        return "C04"
    if 대분류.startswith("루프드레인"):
        return "C05"
    return "C01"


# 대분류별 쉬운 설명 (상품 상세설명·요약설명에 들어감)
GROUP_INFO = {
    "콘크리트-벤치플룸관": ("농수로·배수로용 U형 콘크리트 수로관(벤치플룸관). KS 규격, 광주·전남 현장 직납",
        "벤치플룸관은 바닥이 둥근 U자 모양의 콘크리트 수로관입니다. 물이 바닥에 고이지 않고 잘 흘러가며, 길이 1m 단위로 이어 붙여 시공합니다. "
        "농업용 수로, 도로 옆 배수로, 단지 내 우수 배수에 가장 많이 쓰입니다. 3종은 벽이 두꺼운 일반형, 2종은 벽이 얇고 가벼운 경량형입니다. "
        "규격 표기는 안쪽 폭 × 안쪽 높이 × 벽 두께(mm)이며 단위는 1본(1m)입니다.",
        "벤치플륨관,벤치플룸관,U형수로관,콘크리트수로관,농수로,배수로,광주수로관,전남수로관"),
    "콘크리트-수로뚜껑": ("벤치플룸관·유공수로관 위에 얹는 콘크리트 수로 뚜껑. 길이 2m",
        "콘크리트 수로뚜껑은 수로관 위를 덮는 콘크리트 판입니다. 철제 그레이팅보다 값이 싸고 도난 걱정이 없어 농로·단지 외곽 수로에 많이 씁니다. "
        "규격 뒤의 T 숫자는 맞는 수로관 규격(예: 400T = 400 수로관용)입니다. 길이 2m 기준.",
        "수로뚜껑,콘크리트뚜껑,수로관뚜껑,벤치플륨뚜껑,배수로덮개"),
    "콘크리트-측구수로관": ("도로 가장자리 배수용 측구수로관. 일반형/앵글형, 200~600 규격",
        "측구수로관은 도로 양옆(측구)에 설치해 빗물을 모아 흘려보내는 콘크리트 배수로입니다. 일반형은 윗면에 철제 테두리가 없는 기본형이고, "
        "앵글형은 윗면 양쪽에 ㄱ자 철재(앵글)를 미리 넣어 그레이팅이 흔들리지 않게 한 제품으로 차량 통행 구간에 권장합니다. "
        "규격은 안쪽 폭 × 안쪽 높이 × 벽 두께(mm), 길이 1m. 덮개(그레이팅)는 별도입니다.",
        "측구수로관,측구,도로배수,콘크리트배수로,앵글수로관,측구수로관300,측구수로관400,측구수로관500"),
    "콘크리트-유공수로관": ("벽면에 물구멍이 있는 유공수로관. 땅속 물을 모아 빼는 배수용",
        "유공수로관은 벽면에 작은 구멍이 나 있어 주변 흙 속의 물이 관 안으로 스며들게 만든 수로관입니다. 지하수위가 높은 농지, 옹벽 뒤 배수, "
        "운동장·공원 지하 배수에 씁니다. 규격 표기는 벤치플룸관과 같으며 길이 1m 기준입니다.",
        "유공수로관,유공관,지하배수,농지배수,콘크리트유공관"),
    "콘크리트-보강수로관": ("대형 보강 수로관 1500×1500. 견적 문의",
        "보강수로관은 철근을 넣어 큰 하중을 견디도록 만든 대형 수로관입니다. 현장 조건에 따라 견적을 드립니다.",
        "보강수로관,대형수로관,철근콘크리트수로관"),
    "콘크리트-원형사각수로관": ("겉은 사각, 안은 원형인 원심력 수로관. 매립형(NS02)·돌출형(NS03)·2구형·집수정",
        "원형사각수로관은 겉은 사각형, 안쪽 물길은 원형으로 만든 원심력 콘크리트 수로관입니다. 안이 둥글어 이물질이 덜 걸리고 물이 빠르게 흐릅니다. "
        "NS02(매립형)는 윗면이 지면과 같은 높이로 묻는 제품, NS03(돌출형)은 윗부분이 땅 위로 올라와 경계 역할을 함께 하는 제품입니다. "
        "2구형은 물길이 두 줄이라 배수량이 두 배이고, 집수정은 수로관이 만나거나 꺾이는 지점에 두는 통입니다. 규격은 D(안지름) × 길이(mm).",
        "원형사각수로관,원심력사각수로관,매립형수로관,돌출형수로관,NS-02,NS-03,2구형수로관,집수정"),
    "콘크리트-집수정": ("빗물받이용 콘크리트 집수정. 일반형 / 50앵글형 / 75앵글형, 두께 150T",
        "콘크리트 집수정은 도로·주차장 바닥의 빗물을 모아 배수관으로 보내는 콘크리트 통입니다. 규격은 가로 × 세로 × 높이 × 벽두께(mm)입니다. "
        "일반형은 위에 콘크리트 상판을 따로 얹는 제품(상판 별도), 50앵글형·75앵글형은 윗면에 ㄱ자 철재(앵글)가 들어 있어 그 위에 그레이팅을 바로 얹습니다. "
        "숫자 50/75는 앵글 높이(mm)로, 인도·경하중은 50, 차량 통행 구간은 75를 권장합니다. '앵글형'이라고만 하면 50앵글형으로 준비합니다.",
        "콘크리트집수정,집수정,빗물받이,우수받이,앵글집수정,집수정150T"),
    "콘크리트-맨홀": ("조립식 원형 콘크리트 맨홀 구체(하부·연직·상부). Ø900·1200·1500",
        "조립식 원형맨홀은 하수·우수·전기·통신 관로를 점검하는 맨홀을 공장 제품으로 쌓아 만드는 자재입니다. 맨 아래 '하부구체', 중간에 높이를 맞추는 '연직구체', "
        "뚜껑이 놓이는 '상부구체'를 조합해 필요한 깊이를 만듭니다. 상부구체는 H0 기준 가격이며 높이 100mm마다 추가금이 붙습니다(상세 비고 참고). "
        "규격은 안지름 × 높이 × 벽두께(mm).",
        "콘크리트맨홀,원형맨홀,조립식맨홀,맨홀구체,하부구체,상부구체,연직구체,PC맨홀"),
    "주철-인증맨홀": ("KS 인증 주철 맨홀뚜껑 세트(뚜껑+받침). Ø648·Ø766·Ø918, 1120×620",
        "주철 맨홀뚜껑은 맨홀 위를 덮는 무쇠 뚜껑입니다. '이면도로용'은 골목길·인도처럼 가벼운 하중 구간용, '정중량'과 '회주철'은 차량이 지나는 도로용으로 더 두껍고 무겁습니다. "
        "'세트'는 뚜껑과 받침틀이 함께 들어 있습니다. Ø766·Ø918에 속뚜껑을 포함하면 20,000원이 추가됩니다.",
        "맨홀뚜껑,주철뚜껑,KS맨홀뚜껑,원형맨홀뚜껑,648맨홀뚜껑,사각맨홀뚜껑,이면도로맨홀뚜껑"),
    "주철-원형맨홀": ("Ø300~Ø648 소형 원형 주철 맨홀뚜껑. 오수받이·정화조·차도용",
        "소형 원형 맨홀뚜껑입니다. 오수받이, 정화조 점검구, 소형 점검구에 씁니다. '보도'는 사람만 다니는 곳, '차도'는 차량이 지나는 곳용이며 "
        "'안전망'은 사람이 빠지지 않도록 뚜껑 아래 그물이 달린 제품입니다.",
        "원형맨홀뚜껑,오수받이뚜껑,정화조뚜껑,소형맨홀뚜껑,주철뚜껑"),
    "주철-멀티/칼라": ("멀티맨홀·칼라맨홀·3볼트 잠금 맨홀. 보도용 60t / 차도용 80t",
        "멀티맨홀은 통신·전기 등 여러 용도에 두루 쓰는 주철 맨홀뚜껑이고, 칼라맨홀은 뚜껑 안에 보도블록·타일을 채워 바닥과 같은 모양으로 마감하는 제품입니다. "
        "3볼트 잠금 맨홀은 볼트로 잠가 무단 개방을 막습니다. t 숫자는 뚜껑 두께(mm)로, 보도 60t·차도 80t가 기준입니다.",
        "멀티맨홀,칼라맨홀,잠금맨홀,통신맨홀뚜껑,전기맨홀뚜껑"),
    "주철-이단철개": ("상수도 제수변·지수전용 이단철개(밸브 보호통 뚜껑)",
        "이단철개는 땅속 상수도 밸브(제수변·지수전) 위에 세우는 보호통의 주철 뚜껑입니다. 밸브를 열고 닫을 때 이 뚜껑을 열고 작업합니다.",
        "이단철개,제수변철개,지수전철개,밸브보호통,상수도철개"),
    "주철-인상링": ("맨홀 높이 조절용 인상링(주철·PE). Ø150~Ø918, 사각, 계량기용",
        "인상링은 도로를 덧포장해 높이가 올라갔을 때 기존 맨홀뚜껑을 새 높이에 맞춰 올리는 고리 모양 부속입니다. 맨홀을 다시 만들 필요 없이 링만 끼우면 됩니다. "
        "규격은 맞는 맨홀 뚜껑의 지름(mm)이며, 높이는 비고를 참고하세요.",
        "인상링,맨홀인상링,높이조절링,PE인상링,계량기인상링"),
    "주철-우수통": ("주철제 우수통(빗물받이). 400×500, 400×1000",
        "주철제 우수통은 도로 가장자리에서 빗물을 받는 주철 통입니다. 콘크리트 집수정보다 가볍고 시공이 빠릅니다.",
        "우수통,주철우수통,빗물받이,주철빗물받이"),
    "주철-트렌치": ("차량 통과 시 소음이 없는 무소음 주철 트렌치. 폭 200~500",
        "무소음 트렌치는 주차장 입구, 세차장, 건물 진입로 바닥에 설치하는 주철 배수 덮개입니다. 덮개와 틀이 꼭 맞게 가공돼 차가 지나가도 소리가 나지 않습니다. "
        "규격은 폭 × 길이 × 두께(mm)이며 중차량용 기준 가격입니다(경차량용은 비고 참고).",
        "무소음트렌치,주철트렌치,트렌치덮개,주차장배수,세차장트렌치"),
    "주철-FRP정화조": ("FRP 정화조용 맨홀뚜껑 세트(안전망 포함). 5톤·10톤",
        "FRP 정화조 위에 설치하는 점검구 뚜껑입니다. 사람이 빠지지 않도록 스텐·철근·FRP 재질의 안전망이 함께 들어 있습니다.",
        "정화조맨홀,FRP정화조뚜껑,정화조안전망,정화조점검구"),
    "주철-사각맨홀": ("사각 주철 맨홀뚜껑. 300×1000 ~ 1300×1300, 두께 30/50/70",
        "사각 주철 맨홀뚜껑입니다. 규격은 가로 × 세로 × 두께(mm)이며, 두께 30은 경찰·전기·통신 핸드홀용(뚜껑만), 50은 우수·오수·전기 일반용, 70은 중차량 통행 구간용입니다. "
        "'세트'는 뚜껑과 받침틀 포함, '뚜껑만'은 받침 없이 뚜껑만 드립니다.",
        "사각맨홀뚜껑,핸드홀뚜껑,주철사각뚜껑,전기맨홀뚜껑,통신맨홀뚜껑,우수맨홀뚜껑"),
    "주철-수도계량기": ("수도계량기 보호통용 주철 커버. 13/15, 20/25, 40/50mm",
        "수도계량기 보호통 위에 얹는 주철 커버입니다. 규격은 계량기 구경(mm)입니다.",
        "수도계량기커버,계량기뚜껑,계량기보호통뚜껑"),
    "루프드레인": ("옥상·발코니·교량 배수용 루프드레인(PVC·강관용). 75~200mm",
        "루프드레인은 옥상이나 발코니 바닥에 설치해 빗물을 배수관으로 보내는 배수구 부속입니다. 낙엽이 들어가지 않도록 캡(덮개)이 함께 있습니다. "
        "일반형·주공형·벽체형·나팔형·교량용이 있으며 규격은 연결되는 배수관 지름(mm)입니다.",
        "루프드레인,옥상배수,발코니드레인,PVC드레인,강관드레인,교량드레인"),
    "PE관-직관": ("PE 이중벽관·다중벽관 직관. 150~1000mm, 1·2·3급",
        "PE 이중벽관은 플라스틱(폴리에틸렌) 배수관으로 가볍고 시공이 빠릅니다. 1급이 가장 튼튼하고(차도 매설), 2급·3급 순으로 얇아집니다. 다중(삼중)벽관은 강도가 더 높습니다. "
        "1·2급은 6m, 3급은 4m 길이 기준이며, 유공관은 무공관 단가의 25% 추가입니다. 400만원 이상 주문은 전남권 현장 도착도로 배송합니다.",
        "PE이중벽관,PE관,이중벽관,다중벽관,PE배수관,150mm PE관,300mm PE관"),
    "PE관-부속": ("PE관 연결 부속. PE밴드·스텐밴드·열수축밴드·원봉연결구",
        "PE관을 서로 잇는 부속입니다. PE밴드는 기본 연결용, 스텐(sts)밴드는 조임이 강한 연결용, 열수축밴드는 열을 가해 밀착시키는 방수 연결용, 원봉연결구는 소켓형 연결구입니다. "
        "규격은 맞는 관의 지름(mm)입니다.",
        "PE밴드,열수축밴드,스텐밴드,원봉연결구,PE관부속"),
    "PE관-엘보": ("PE 이중벽관 엘보(22.5·45도 / 90도). 150~800mm",
        "엘보는 관의 방향을 꺾을 때 쓰는 부속입니다. 22.5도·45도는 완만하게, 90도는 직각으로 꺾습니다. 길이 20cm 미만 기준이며 별도 길이 주문 시 금액이 달라집니다.",
        "PE엘보,이중벽관엘보,PE관엘보,90도엘보,45도엘보"),
    "PE관-이경관": ("PE 이중벽관 이경관(구경이 다른 관 연결). 150~600mm",
        "이경관은 지름이 다른 두 관을 잇는 부속입니다. 규격은 큰 관 지름 × 작은 관 지름(mm)이며, 같은 지름끼리는 분기(T) 연결용입니다. 길이 1m 미만 기준.",
        "PE이경관,이경관,PE관연결,분기관"),
}


def fmt_won(v):
    return f"{int(v):,}원"


def sale_price(판매단가):
    """카페24 판매가 (부가세 포함, 10원 단위 반올림)."""
    if not VAT_INCLUDED:
        return int(판매단가)
    return int(round(판매단가 * 1.1 / 10.0) * 10)


def parse_note(비고):
    """비고에서 나라장터 식별번호·중량을 뽑고 나머지를 돌려준다."""
    ident = ""
    weight = ""
    m = re.search(r"식별번호\s*(\d+)", 비고)
    if m:
        ident = m.group(1)
    m = re.search(r"중량\s*([\d,\.]+)\s*kg", 비고)
    if m:
        weight = m.group(1).replace(",", "")
    rest = re.sub(r"식별번호\s*\d+\s*·?\s*", "", 비고)
    rest = re.sub(r"·?\s*중량\s*[\d,\.]+\s*kg", "", rest).strip(" ·")
    return ident, weight, rest


def load_items():
    with open(ITEMS_CSV, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    items = []
    for r in rows:
        try:
            price = int(float(r["판매단가"] or 0))
        except ValueError:
            price = 0
        ident, weight, note = parse_note(r.get("비고", "") or "")
        items.append(dict(
            code=r["코드"].strip(), group=r["대분류"].strip(), name=r["품목명"].strip(), spec=r["규격"].strip(),
            unit=(r["단위"] or "개").strip(), price=price, ident=ident, weight=weight, note=note,
        ))
    return items


def family_of(item):
    """옵션형 등록용 제품군 이름: 품목명 끝의 규격 토큰(숫자 포함)을 떼어 낸다."""
    toks = item["name"].split(" ")
    if len(toks) > 1 and re.search(r"\d", toks[-1]) and not toks[-1].endswith("형"):
        return " ".join(toks[:-1]), toks[-1]
    return item["name"], ""


def detail_html(group, title, rows):
    """상품 상세설명 HTML. rows = [(규격표시, 단위, 판매단가, ident, note)]"""
    esc = html.escape
    summary, desc, _ = GROUP_INFO.get(group, ("", "", ""))
    tr = ""
    for spec, unit, price, ident, note in rows:
        if price == 0:
            p_txt = "견적 문의"
            p_ex = "-"
        else:
            p_txt = fmt_won(sale_price(price)) if VAT_INCLUDED else fmt_won(price)
            p_ex = fmt_won(price)
        tr += ("<tr>"
               f"<td style='border:1px solid #ccc;padding:8px'>{esc(spec)}</td>"
               f"<td style='border:1px solid #ccc;padding:8px;text-align:center'>{esc(unit)}</td>"
               f"<td style='border:1px solid #ccc;padding:8px;text-align:right'>{p_txt}</td>"
               f"<td style='border:1px solid #ccc;padding:8px;text-align:right'>{p_ex}</td>"
               f"<td style='border:1px solid #ccc;padding:8px;text-align:center'>{esc(ident)}</td>"
               f"<td style='border:1px solid #ccc;padding:8px;font-size:0.9em'>{esc(note)}</td>"
               "</tr>")
    price_head = "판매가(부가세 포함)" if VAT_INCLUDED else "판매가(부가세 별도)"
    return (
        "<div style='max-width:900px;margin:0 auto;font-family:\"Malgun Gothic\",\"Apple SD Gothic Neo\",sans-serif;line-height:1.7;color:#222'>"
        f"<p style='text-align:center'><img src='{LOGO}' alt='{COMPANY}' style='max-width:180px'></p>"
        f"<h2 style='border-left:6px solid #1D5C57;padding-left:12px'>{esc(title)}</h2>"
        f"<p><b>{esc(summary)}</b></p>"
        f"<p>{esc(desc)}</p>"
        "<h3>규격 및 단가</h3>"
        "<div style='overflow-x:auto'><table style='border-collapse:collapse;width:100%'>"
        "<thead><tr style='background:#D6E4E1'>"
        "<th style='border:1px solid #ccc;padding:8px'>규격</th><th style='border:1px solid #ccc;padding:8px'>단위</th>"
        f"<th style='border:1px solid #ccc;padding:8px'>{price_head}</th><th style='border:1px solid #ccc;padding:8px'>부가세 별도 단가</th>"
        "<th style='border:1px solid #ccc;padding:8px'>나라장터 식별번호</th><th style='border:1px solid #ccc;padding:8px'>비고</th>"
        f"</tr></thead><tbody>{tr}</tbody></table></div>"
        "<p style='font-size:0.9em;color:#555'>※ 단가는 공장 상차도 기준이며 운반비는 별도(착불)입니다. 하차는 현장 부담입니다. "
        "KS 규격품은 시험성적서·자재승인서류를 납품 시 함께 드립니다. 대량 주문은 별도 견적을 드립니다.</p>"
        "<h3>배송 안내</h3>"
        f"<p>{esc(SHIPPING_GUIDE)}</p>"
        "<h3>교환·반품</h3>"
        f"<p>{esc(RETURN_GUIDE)}</p>"
        "<h3>문의</h3>"
        f"<p>{esc(COMPANY)}<br>전화 {PHONE} (평일 08:00~18:00, 주말·공휴일 휴무)<br>팩스 {FAX}<br>이메일 {EMAIL}<br>"
        f"{esc(ADDR)}<br><a href='{HOMEPAGE}'>{HOMEPAGE}</a></p>"
        "</div>"
    )


def base_row(group, name, own_code, price, desc_html, model, weight=""):
    """카페24 열 순서대로 값 채우기. 비워둔 열은 관리자 기본값 사용."""
    summary, _, keywords = GROUP_INFO.get(group, ("", "", ""))
    r = {c: "" for c in CAFE24_COLUMNS}
    quote = price == 0
    sp = sale_price(price) if not quote else 0
    r.update({
        "상품코드": "",                      # 신규 등록 시 비움 (자동 발급)
        "자체 상품코드": own_code,            # data/품목.csv 의 코드 그대로 → 견적·재고와 같은 코드
        "진열상태": "T",
        "판매상태": "T",
        "상품분류 번호": "",                 # 관리자에서 분류 만든 뒤 번호 입력 (README 참고)
        "상품명": name,
        "상품명(관리용)": name,
        "모델명": model,
        "상품 요약설명": summary,
        "상품 간략설명": summary,
        "상품 상세설명": desc_html,
        "모바일 상품 상세설명 설정": "F",     # F = PC 상세설명을 모바일에도 사용
        "검색어설정": keywords,
        "과세구분": "A",                      # A=과세상품
        "소비자가": "" if quote else sp,
        "공급가": "" if quote else price,     # 부가세 별도 단가
        "판매가": sp,
        "판매가 대체문구 사용": "T" if quote else "F",
        "판매가 대체문구": "견적문의" if quote else "",
        "주문수량 제한 기준": "O",            # O=주문 기준
        "최소 주문수량(이상)": 1,
        "적립금": 0,
        "적립금 구분": "W",                   # W=원
        "성인인증": "F",
        "옵션사용": "F",
        "추가입력옵션": "T",
        "추가입력옵션 명칭": "납품 현장 주소·하차장비 유무",
        "추가입력옵션 선택/필수여부": "F",     # F=선택
        "입력글자수(자)": 100,
        "제조사": COMPANY,
        "공급사": COMPANY,
        "브랜드": "현담토목",
        "자체분류 코드": category_of(group),
        "원산지": "국내산",
        "상품결제안내": PAYMENT_GUIDE,
        "상품배송안내": SHIPPING_GUIDE,
        "교환/반품안내": RETURN_GUIDE,
        "서비스문의/안내": SERVICE_GUIDE,
        "배송정보": "F",                      # F=기본 배송설정 사용 (관리자 배송설정에서 화물/착불 지정)
        "국내/해외배송": "A",                 # A=국내
        "배송지역": "전국지역",
        "스토어픽업 설정": "T",               # 공장 직접 수령 가능
        "상품 전체중량(kg)": weight,
        "검색엔진 노출 설정": "T",
        "검색엔진 노출 제목": f"{name} | {COMPANY} 광주 직접생산",
        "검색엔진 노출 작성자": COMPANY,
        "검색엔진 노출 설명": summary,
        "검색엔진 노출 키워드": keywords,
        "상품 유효기간 사용": "F",
    })
    return r


def style_header(ws, ncol):
    fill = PatternFill("solid", fgColor="D6E4E1")
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def write_rows(ws, rows):
    ws.append(CAFE24_COLUMNS)
    for r in rows:
        ws.append([r[c] for c in CAFE24_COLUMNS])
    style_header(ws, len(CAFE24_COLUMNS))
    widths = {"상품명": 46, "상품 요약설명": 40, "상품 간략설명": 40, "상품 상세설명": 30, "검색어설정": 40,
              "자체 상품코드": 16, "모델명": 26, "상품결제안내": 30, "상품배송안내": 30, "교환/반품안내": 30,
              "서비스문의/안내": 30, "옵션 입력": 60, "검색엔진 노출 제목": 40, "검색엔진 노출 설명": 40, "검색엔진 노출 키워드": 40}
    for i, c in enumerate(CAFE24_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)


def main():
    items = load_items()
    missing = sorted({it["group"] for it in items} - set(GROUP_INFO))
    if missing:
        print("경고: 설명이 없는 대분류 →", missing, file=sys.stderr)

    wb = Workbook()

    # 1) 품목별 개별 상품 (권장)
    ws1 = wb.active
    ws1.title = "카페24_규격별등록"
    rows1 = []
    for it in items:
        name = f"{it['name']} {it['spec']}".strip()
        rows = [(it["spec"], it["unit"], it["price"], it["ident"], it["note"])]
        rows1.append(base_row(it["group"], name, it["code"], it["price"],
                              detail_html(it["group"], name, rows), it["spec"], it["weight"]))
    write_rows(ws1, rows1)

    # 2) 옵션형 (제품군 1개 = 상품 1개)
    families = {}
    order = []
    for it in items:
        fam, tail = family_of(it)
        key = (it["group"], fam)
        if key not in families:
            families[key] = []
            order.append(key)
        families[key].append((it, tail))
    ws2 = wb.create_sheet("카페24_옵션형등록")
    rows2 = []
    addprice_rows = []
    for key in order:
        group, fam = key
        members = families[key]
        first_price = members[0][0]["price"]
        opt_values = []
        rows = []
        for it, tail in members:
            label = f"{tail} {it['spec']}".strip() if tail else it["spec"]
            label = label.replace("|", "/").replace("{", "(").replace("}", ")")
            opt_values.append(label)
            rows.append((label, it["unit"], it["price"], it["ident"], it["note"]))
            addprice_rows.append([members[0][0]["code"], fam, it["code"], label, sale_price(it["price"]) if it["price"] else 0,
                                  (sale_price(it["price"]) - sale_price(first_price)) if it["price"] and first_price else 0])
        fam_code = "F-" + members[0][0]["code"]
        r = base_row(group, fam, fam_code, first_price, detail_html(group, fam, rows), fam)
        r.update({
            "옵션사용": "T",
            "옵션구성방식": "C",                 # C=조합 일체선택형
            "옵션 표시방식": "S",                # S=셀렉트박스
            "옵션 입력": "규격{" + "|".join(opt_values) + "}",
            "필수여부": "T",
        })
        rows2.append(r)
    write_rows(ws2, rows2)

    # 3) 옵션 추가금액표
    ws3 = wb.create_sheet("옵션추가금액표")
    ws3.append(["옵션형 자체 상품코드(대표 품목코드)", "제품군", "품목코드", "규격(옵션값)", "규격 판매가", "기본가 대비 추가금액 (옵션형 등록 시 입력)"])
    for r in addprice_rows:
        ws3.append(r)
    style_header(ws3, 6)
    for i, w in enumerate([22, 34, 12, 44, 14, 30], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # 4) 제품 마스터 단가표
    ws4 = wb.create_sheet("제품마스터_단가표")
    ws4.append(["카페24 분류", "대분류", "품목코드", "품목명", "규격", "단위", "판매단가(부가세 별도)", "카페24 판매가(부가세 포함)", "나라장터 식별번호", "중량(kg)", "비고"])
    for it in items:
        ws4.append([CAT_NAME[category_of(it["group"])], it["group"], it["code"], it["name"], it["spec"], it["unit"],
                    it["price"] if it["price"] else "견적문의", sale_price(it["price"]) if it["price"] else "견적문의",
                    it["ident"], it["weight"], it["note"]])
    style_header(ws4, 11)
    for i, w in enumerate([22, 20, 10, 34, 24, 6, 18, 20, 16, 10, 50], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # 5) 열 설명
    ws5 = wb.create_sheet("카페24_열설명")
    ws5.append(["열 이름", "이 파일에 넣은 값", "설명"])
    notes = [
        ("상품코드", "비움", "신규 등록은 비워 둡니다. 카페24가 자동으로 만들어 줍니다. (수정 업로드 때만 입력)"),
        ("자체 상품코드", "data/품목.csv 의 품목코드", "견적·재고 시스템과 같은 코드. 나중에 가격 수정 엑셀을 올릴 때 이 코드로 찾습니다."),
        ("진열상태 / 판매상태", "T", "T=진열함/판매함, F=안 함"),
        ("상품분류 번호", "비움", "관리자 > 상품 > 분류 관리에서 분류를 먼저 만들면 번호가 생깁니다. 그 번호를 넣으면 자동 분류됩니다. 비워도 등록은 됩니다."),
        ("상품명", "품목명 + 규격", "고객에게 보이는 이름"),
        ("상품 상세설명", "HTML", "제품 설명·규격표·배송안내가 들어간 HTML. 업로드 후 이미지만 추가하면 됩니다."),
        ("검색어설정", "쉼표로 구분된 검색어", "쇼핑몰 안 검색과 네이버 노출용"),
        ("과세구분", "A", "A=과세상품, B=면세, C=영세"),
        ("판매가", "판매단가 × 1.1 (부가세 포함)" if VAT_INCLUDED else "판매단가 (부가세 별도)", "data/품목.csv 판매단가 기준. 견적문의 상품은 0"),
        ("공급가", "판매단가 (부가세 별도)", "견적서에 쓰는 부가세 별도 단가를 그대로 기록"),
        ("판매가 대체문구 사용 / 문구", "T + 견적문의", "가격 대신 '견적문의' 글자를 보여줍니다(단가 없는 상품만)"),
        ("주문수량 제한 기준", "O", "O=주문 기준, P=품목 기준"),
        ("적립금 구분", "W", "W=원 단위, P=퍼센트"),
        ("옵션사용", "F 또는 T", "규격별등록 시트는 F(옵션 없음), 옵션형 시트는 T"),
        ("옵션구성방식", "C", "C=조합 일체선택형, S=조합 분리선택형, E=상품 연동형, F=독립 선택형"),
        ("옵션 입력", "규격{값1|값2|…}", "옵션명{옵션값|옵션값}. 옵션이 여럿이면 // 로 구분"),
        ("추가입력옵션", "T", "고객이 주문할 때 '납품 현장 주소·하차장비 유무'를 적을 수 있는 칸"),
        ("이미지등록(상세) 등", "비움", "카페24 웹FTP에 사진을 올린 뒤 파일 이름만 적습니다. 사진이 준비되면 상품 수정 화면에서 올려도 됩니다."),
        ("배송정보", "F", "F=쇼핑몰 기본 배송 설정 사용. 관리자 > 설정 > 배송 설정에서 '화물배송/착불'을 기본으로 지정해 두세요."),
        ("스토어픽업 설정", "T", "공장에서 직접 가져가는 손님을 위해 켬"),
        ("상품 전체중량(kg)", "비고의 중량", "콘크리트 제품은 품목표의 중량을 넣었습니다(운반 계획용)"),
        ("검색엔진 노출 …", "제목/설명/키워드", "네이버·구글 검색 노출용 문구"),
    ]
    for n in notes:
        ws5.append(list(n))
    style_header(ws5, 3)
    for i, w in enumerate([26, 30, 90], 1):
        ws5.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT_XLSX)
    print(f"saved {OUT_XLSX}: {len(rows1)} 규격별 상품, {len(rows2)} 옵션형 상품 (부가세 {'포함' if VAT_INCLUDED else '별도'} 판매가)")


if __name__ == "__main__":
    main()
